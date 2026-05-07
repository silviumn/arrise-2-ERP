"""
Vizofia — procesare comenzi lentile contact (Arrise → Neomanager).
Logica de transformare a tabelului de comandă în formatul cerut de furnizor.
"""
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# ============================================================================
# CONFIGURARE: rădăcini coduri produse
# ============================================================================

# Lentile sferice: rădăcină + dioptrie semnată cu 4 cifre (ex: -0150)
SFERIC_ROOTS = {
    'biofinity': 'LBIOFSH386',
    'avaira vitality': 'LAVVITY384',
    'clariti 1 day': 'LCLARDSI86',
    'energys': 'LBFNTBE386',
}

# Lentile torice: rădăcină + sferă (4 cifre cu semn) + cilindru (3 cifre fără semn) + ax (3 cifre)
TORIC_ROOTS = {
    'biofinity toric': 'LBFNTRC387',
    'clariti 1 day toric': 'LCLARDTI86',
    'avaira vitality toric': 'LAVARVT385',
    'proclear toric': 'LPCCRTR388',
}

# Variante speciale
SPECIAL_ROOTS = {
    'biofinity_6': 'LBIOFSH686',         # 6-pack Biofinity (aceeași dioptrie ambii ochi, gama standard)
    'biofinity_xr': 'LBFXRSH386',        # Biofinity XR sferic (+8.50..+15.50 sau -20.00..-12.50)
    'biofinity toric xr': 'LBFXRTR387',  # Biofinity Toric XR (CYL magnitudine > 2.25)
}

# Fallback toric → sferic când valoarea ochiului are DOAR sfera (fără CYL/AX).
TORIC_TO_SFERIC = {
    'biofinity toric': 'biofinity',
    'avaira vitality toric': 'avaira vitality',
    'clariti 1 day toric': 'clariti 1 day',
}

# Praguri XR Biofinity sferic
BIOFINITY_XR_POS_MIN = 8.50
BIOFINITY_XR_POS_MAX = 15.50
BIOFINITY_XR_NEG_MIN = -20.00
BIOFINITY_XR_NEG_MAX = -12.50
BIOFINITY_TORIC_XR_CYL_THRESHOLD = 2.25  # strict mai mare → XR


# ============================================================================
# PARSARE
# ============================================================================

def fmt_sphere(value):
    """SF-1.50 → '-0150'. SF-1 → '-0100'. SF+0.75 → '+0075'."""
    if value is None:
        return None
    s = str(value).strip()
    m = re.search(r'([+-])\s*(\d+)(?:\.(\d+))?', s)
    if not m:
        return None
    sign = m.group(1)
    int_part = int(m.group(2))
    dec_str = ((m.group(3) or '') + '00')[:2]
    return f"{sign}{int_part:02d}{dec_str}"


def fmt_toric(value):
    """SF-5.00 CYL-1.25 AX 180 → '-0500125180'.
    Plan în toric → sferă=0000 (fără semn)."""
    if value is None:
        return None
    s = str(value).strip()
    is_plan_value = 'plan' in s.lower()
    m_sf = re.search(r'SF\s*([+-])\s*(\d+)(?:\.(\d+))?', s, re.IGNORECASE)
    m_cyl = re.search(r'CYL\s*[+-]?\s*(\d+)(?:\.(\d+))?', s, re.IGNORECASE)
    m_ax = re.search(r'AX\s*(\d+)', s, re.IGNORECASE)

    if m_sf and not is_plan_value:
        sign = m_sf.group(1)
        sf_int = int(m_sf.group(2))
        sf_dec = ((m_sf.group(3) or '') + '00')[:2]
        sphere = f"{sign}{sf_int:02d}{sf_dec}"
    else:
        sphere = "0000"

    if m_cyl:
        cyl_int = m_cyl.group(1)
        cyl_dec = ((m_cyl.group(2) or '') + '00')[:2]
        cyl = f"{cyl_int}{cyl_dec}"
    else:
        cyl = "000"

    if m_ax:
        ax = f"{int(m_ax.group(1)):03d}"
    else:
        ax = "000"

    return f"{sphere}{cyl}{ax}"


def extract_sphere_value(value):
    """SF-1.50 → -1.5 (float)."""
    if value is None:
        return None
    s = str(value).strip()
    m = re.search(r'([+-])\s*(\d+)(?:\.(\d+))?', s)
    if not m:
        return None
    sign = 1 if m.group(1) == '+' else -1
    int_part = int(m.group(2))
    if m.group(3):
        dec_part = (m.group(3) + '00')[:2]
    else:
        dec_part = '0'
    return sign * float(f"{int_part}.{dec_part}")


def extract_cyl_magnitude(value):
    """Magnitudinea CYL ca float (ex: 1.25, 2.50). 0.0 dacă lipsește."""
    if value is None:
        return 0.0
    s = str(value).strip()
    m = re.search(r'CYL\s*[+-]?\s*(\d+)(?:\.(\d+))?', s, re.IGNORECASE)
    if not m:
        return 0.0
    int_part = int(m.group(1))
    dec_part = (m.group(2) or '0')
    dec_part = (dec_part + '00')[:2]
    return float(f"{int_part}.{dec_part}")


def is_xr_sphere(value):
    """True dacă dioptria sferică Biofinity este în gama XR."""
    sphere = extract_sphere_value(value)
    if sphere is None:
        return False
    if BIOFINITY_XR_POS_MIN <= sphere <= BIOFINITY_XR_POS_MAX:
        return True
    if BIOFINITY_XR_NEG_MIN <= sphere <= BIOFINITY_XR_NEG_MAX:
        return True
    return False


def is_plan(value):
    """True dacă valoarea este exact 'Plan'/'PLAN'."""
    if value is None:
        return True
    return str(value).strip().lower() == 'plan'


def is_sferic_only(value):
    """True dacă valoarea conține doar sferă (fără CYL și fără AX)."""
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    has_cyl = re.search(r'CYL', s, re.IGNORECASE) is not None
    has_ax = re.search(r'AX', s, re.IGNORECASE) is not None
    return not (has_cyl or has_ax)


# ============================================================================
# CONSTRUCȚIE COD
# ============================================================================

def build_code(product_type, eye_value):
    """Construiește codul produs complet (rădăcină + dioptrie, fără spațiu)."""
    pt = product_type.strip().lower()

    # Fallback: toric cu doar sferă → sferic echivalent
    if pt in TORIC_TO_SFERIC and is_sferic_only(eye_value):
        pt = TORIC_TO_SFERIC[pt]

    if pt in TORIC_ROOTS:
        formatted = fmt_toric(eye_value)
        if formatted is None:
            return None
        if pt == 'biofinity toric':
            cyl_mag = extract_cyl_magnitude(eye_value)
            root = SPECIAL_ROOTS['biofinity toric xr'] if cyl_mag > BIOFINITY_TORIC_XR_CYL_THRESHOLD else TORIC_ROOTS[pt]
        else:
            root = TORIC_ROOTS[pt]
        return f"{root}{formatted}"

    if pt in SFERIC_ROOTS:
        formatted = fmt_sphere(eye_value)
        if formatted is None:
            return None
        if pt == 'biofinity' and is_xr_sphere(eye_value):
            root = SPECIAL_ROOTS['biofinity_xr']
        else:
            root = SFERIC_ROOTS[pt]
        return f"{root}{formatted}"

    return None


# ============================================================================
# PROCESARE FIȘIER
# ============================================================================

def _find_col(header_row, *patterns):
    """Returnează numărul coloanei (1-based) unde un header se potrivește cu unul
    dintre pattern-urile regex date (case-insensitive, cu/fără diacritice).
    Pattern-urile sunt încercate în ordine — primul match câștigă."""
    for pattern in patterns:
        for cell in header_row:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            if re.search(pattern, text, re.IGNORECASE):
                return cell.column
    return None


def process_file(input_path, output_path):
    """Procesează un fișier Excel de comenzi. Returnează statistici."""
    wb_in = load_workbook(input_path)
    ws_in = wb_in.active

    # ------------------------------------------------------------------
    # Detectare coloane după header (robust la schimbarea ordinii)
    # ------------------------------------------------------------------
    header_row = list(ws_in[1])

    col_name     = _find_col(header_row, r'\bnume\b', r'pacient', r'client')
    col_type     = _find_col(header_row, r'tip\s*lentil[ăa]', r'\btip\b')
    col_od       = _find_col(header_row, r'od\s*lentil[ăa]', r'\bod\b')
    col_os       = _find_col(header_row, r'os\s*lentil[ăa]', r'\bos\b')
    col_delivery = _find_col(header_row, r'livrare', r'termen')

    missing = []
    if col_name is None:     missing.append('Nume')
    if col_type is None:     missing.append('Tip lentile')
    if col_od is None:       missing.append('OD')
    if col_os is None:       missing.append('OS')
    if col_delivery is None: missing.append('Livrare')

    if missing:
        # Construiește un mesaj de eroare util care arată ce s-a găsit
        found_headers = [str(c.value).strip() for c in header_row if c.value is not None]
        raise ValueError(
            f"Nu am putut identifica următoarele coloane în fișier: {', '.join(missing)}. "
            f"Headere găsite: {found_headers}. "
            f"Verifică că fișierul conține coloane pentru Nume, Tip lentile, OD, OS și Livrare."
        )

    output_rows = [['Observatii', 'Cod Produs', 'Cantitate', 'Comanda livrare']]
    skipped_plan = 0
    errors = []
    unknown_products = set()

    for row_idx in range(2, ws_in.max_row + 1):
        name = ws_in.cell(row=row_idx, column=col_name).value
        product_type = ws_in.cell(row=row_idx, column=col_type).value
        od = ws_in.cell(row=row_idx, column=col_od).value
        os_val = ws_in.cell(row=row_idx, column=col_os).value
        delivery = ws_in.cell(row=row_idx, column=col_delivery).value

        if name is None or product_type is None:
            continue
        name = re.sub(r'\s+', ' ', str(name).strip())

        pt_lower = product_type.strip().lower()

        if pt_lower not in TORIC_ROOTS and pt_lower not in SFERIC_ROOTS:
            unknown_products.add(product_type)
            errors.append(f"Rând {row_idx} ({name}): tip de lentilă necunoscut: {product_type!r}")
            continue

        def eff_pt(eye_value):
            if pt_lower in TORIC_TO_SFERIC and is_sferic_only(eye_value):
                return TORIC_TO_SFERIC[pt_lower]
            return pt_lower

        od_eff = eff_pt(od)
        os_eff = eff_pt(os_val)

        # AO merge: doar pentru Biofinity sferic (NU Energys, NU alte sferice)
        if od_eff == 'biofinity' and os_eff == 'biofinity' \
                and not is_plan(od) and not is_plan(os_val):
            od_fmt = fmt_sphere(od)
            os_fmt = fmt_sphere(os_val)
            if od_fmt is not None and od_fmt == os_fmt:
                if is_xr_sphere(od):
                    code = f"{SPECIAL_ROOTS['biofinity_xr']}{od_fmt}"
                else:
                    code = f"{SPECIAL_ROOTS['biofinity_6']}{od_fmt}"
                output_rows.append([f"AO {name}", code, 1, delivery])
                continue
            elif od_fmt is None or os_fmt is None:
                errors.append(f"Rând {row_idx} ({name}) Biofinity: nu pot parsa OD={od!r} sau OS={os_val!r}")
                continue

        # Logică pe ochi:
        # - "Plan" singur (fără CYL/AX) → se sare rândul, indiferent de produs.
        #   (un toric fără CYL/AX nu are sens; pentru sferic, e oricum regula veche)
        # - "Plan CYL-X AX Y" la toric → fmt_toric pune sferă=0000, rândul rămâne
        skip_od = is_plan(od)
        if not skip_od:
            code = build_code(product_type, od)
            if code is None:
                errors.append(f"Rând {row_idx} ({name}) OD: nu pot construi codul pentru {od!r}")
            else:
                output_rows.append([f"OD {name}", code, 1, delivery])
        else:
            skipped_plan += 1

        skip_os = is_plan(os_val)
        if not skip_os:
            code = build_code(product_type, os_val)
            if code is None:
                errors.append(f"Rând {row_idx} ({name}) OS: nu pot construi codul pentru {os_val!r}")
            else:
                output_rows.append([f"OS {name}", code, 1, delivery])
        else:
            skipped_plan += 1

    # Scriere output
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = 'Sheet1'
    for r_idx, row in enumerate(output_rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws_out.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name='Calibri', size=11, bold=(r_idx == 1))

    ws_out.column_dimensions['A'].width = 35
    ws_out.column_dimensions['B'].width = 25
    ws_out.column_dimensions['C'].width = 12
    ws_out.column_dimensions['D'].width = 45

    wb_out.save(output_path)

    return {
        'total_rows': len(output_rows) - 1,
        'skipped_plan': skipped_plan,
        'errors': errors,
        'unknown_products': sorted(unknown_products),
    }
