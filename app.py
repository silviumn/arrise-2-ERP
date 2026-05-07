"""
Conversie comenzi lentile (Arrise → Neomanager).
Design: minimal modern · Inter typography · indigo accent · stone neutrals.
"""
import tempfile
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from processor import process_file, SFERIC_ROOTS, TORIC_ROOTS


# ============================================================================
# Configurare pagină
# ============================================================================
st.set_page_config(
    page_title="Conversie comenzi lentile",
    page_icon="◐",
    layout="centered",
    initial_sidebar_state="collapsed",
)


# ============================================================================
# Design system — CSS
# ============================================================================
st.markdown("""
<style>
    @import url('https://rsms.me/inter/inter.css');

    /* ---- Tokens ---- */
    :root {
        --bg: #FAFAF9;
        --surface: #FFFFFF;
        --surface-muted: #F5F5F4;
        --border: #E7E5E4;
        --border-strong: #D6D3D1;
        --text: #1C1917;
        --text-2: #44403C;
        --text-muted: #78716C;
        --text-subtle: #A8A29E;

        --accent: #4F46E5;
        --accent-hover: #4338CA;
        --accent-active: #3730A3;
        --accent-soft: #EEF2FF;
        --accent-soft-2: #E0E7FF;
        --accent-text: #312E81;

        --success: #059669;
        --success-soft: #ECFDF5;
        --warning: #D97706;
        --warning-soft: #FFFBEB;
        --error: #DC2626;
        --error-soft: #FEF2F2;

        --shadow-sm: 0 1px 2px rgba(28, 25, 23, 0.04);
        --shadow-md: 0 1px 3px rgba(28, 25, 23, 0.05), 0 4px 12px -4px rgba(28, 25, 23, 0.06);
        --shadow-lg: 0 4px 16px -4px rgba(79, 70, 229, 0.18), 0 2px 6px rgba(28, 25, 23, 0.06);

        --radius-sm: 8px;
        --radius-md: 12px;
        --radius-lg: 16px;
    }

    /* ---- Reset ---- */
    .stApp {
        background: var(--bg);
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, system-ui, sans-serif;
        font-feature-settings: 'cv02', 'cv03', 'cv04', 'cv11';
    }
    header[data-testid="stHeader"] { background: transparent; height: 0; }
    #MainMenu, footer, [data-testid="stToolbar"] { display: none; }

    .block-container {
        padding: 2.75rem 1.25rem 4rem;
        max-width: 720px;
    }

    /* ---- Typography ---- */
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Inter', sans-serif;
        color: var(--text);
        letter-spacing: -0.022em;
        font-feature-settings: 'cv02', 'cv03', 'cv04', 'cv11';
    }
    p, .stMarkdown { color: var(--text-2); font-family: 'Inter', sans-serif; }
    code {
        font-family: 'JetBrains Mono', ui-monospace, 'SF Mono', Menlo, monospace;
        background: var(--surface-muted);
        padding: 0.12em 0.4em;
        border-radius: 4px;
        font-size: 0.86em;
        color: var(--text);
        border: 1px solid var(--border);
    }

    /* ---- Hero ---- */
    .hero {
        margin-bottom: 1.75rem;
    }
    .logo-row {
        display: flex;
        align-items: center;
        gap: 0.7rem;
        margin-bottom: 1.4rem;
    }
    .logo-mark {
        width: 32px;
        height: 32px;
        border-radius: 9px;
        background: linear-gradient(135deg, var(--accent) 0%, var(--accent-hover) 100%);
        display: flex;
        align-items: center;
        justify-content: center;
        flex-shrink: 0;
        box-shadow: 0 2px 8px -2px rgba(79, 70, 229, 0.4);
    }
    .logo-mark::after {
        content: '';
        width: 14px;
        height: 14px;
        border: 2px solid white;
        border-radius: 50%;
        border-right-color: transparent;
        border-top-color: transparent;
    }
    .logo-text {
        font-size: 0.9rem;
        font-weight: 600;
        color: var(--text-muted);
        letter-spacing: -0.01em;
    }
    h1.app-title {
        font-size: 2rem;
        font-weight: 700;
        color: var(--text);
        letter-spacing: -0.035em;
        line-height: 1.1;
        margin: 0 0 0.55rem;
    }
    .app-subtitle {
        font-size: 1rem;
        font-weight: 400;
        color: var(--text-muted);
        line-height: 1.5;
        max-width: 540px;
    }
    .app-subtitle .pill {
        display: inline-block;
        padding: 0.1rem 0.5rem;
        background: var(--accent-soft);
        color: var(--accent-text);
        border-radius: 5px;
        font-weight: 500;
        font-size: 0.92em;
        font-family: 'JetBrains Mono', ui-monospace, monospace;
    }

    /* ---- Section labels ---- */
    .section-label {
        display: flex;
        align-items: center;
        gap: 0.6rem;
        margin: 1.75rem 0 0.85rem;
    }
    .section-num {
        font-size: 0.68rem;
        font-weight: 700;
        color: var(--text-subtle);
        letter-spacing: 0.14em;
        font-variant-numeric: tabular-nums;
    }
    .section-title {
        font-size: 0.92rem;
        font-weight: 600;
        color: var(--text);
        letter-spacing: -0.005em;
    }
    .section-line {
        flex: 1;
        height: 1px;
        background: var(--border);
        margin-left: 0.4rem;
    }

    /* ---- File uploader ---- */
    [data-testid="stFileUploader"] {
        background: var(--surface);
        border: 1.5px dashed var(--border-strong);
        border-radius: var(--radius-md);
        padding: 0.85rem;
        transition: all 0.18s ease;
    }
    [data-testid="stFileUploader"]:hover {
        border-color: var(--accent);
        background: var(--accent-soft);
    }
    [data-testid="stFileUploader"] section,
    [data-testid="stFileUploaderDropzone"] {
        background: transparent !important;
        border: none !important;
        padding: 0.6rem !important;
    }
    [data-testid="stFileUploader"] section button,
    [data-testid="stFileUploaderDropzone"] button {
        background: var(--surface) !important;
        color: var(--accent) !important;
        border: 1px solid var(--accent) !important;
        border-radius: var(--radius-sm) !important;
        font-weight: 500 !important;
        padding: 0.45rem 1rem !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.86rem !important;
        transition: all 0.15s ease;
        box-shadow: var(--shadow-sm);
    }
    [data-testid="stFileUploader"] section button:hover,
    [data-testid="stFileUploaderDropzone"] button:hover {
        background: var(--accent) !important;
        color: white !important;
    }
    [data-testid="stFileUploader"] section small,
    [data-testid="stFileUploader"] section span,
    [data-testid="stFileUploaderDropzone"] small,
    [data-testid="stFileUploaderDropzone"] span {
        color: var(--text-muted) !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.86rem !important;
    }

    /* ---- Lens type pills ---- */
    .pills-section { margin-top: 1.25rem; }
    .pills-label {
        font-size: 0.68rem;
        font-weight: 700;
        color: var(--text-subtle);
        text-transform: uppercase;
        letter-spacing: 0.12em;
        margin-bottom: 0.55rem;
    }
    .pills {
        display: flex;
        flex-wrap: wrap;
        gap: 0.4rem;
    }
    .pill-item {
        background: var(--surface);
        color: var(--text-2);
        padding: 0.32rem 0.72rem;
        border-radius: 6px;
        font-size: 0.8rem;
        font-weight: 500;
        border: 1px solid var(--border);
    }
    .pill-item.toric {
        background: var(--accent-soft);
        color: var(--accent-text);
        border-color: var(--accent-soft-2);
    }

    /* ---- File chip ---- */
    .file-chip {
        display: flex;
        align-items: center;
        gap: 0.7rem;
        padding: 0.7rem 0.85rem;
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius-md);
        margin-bottom: 1rem;
    }
    .file-chip-ic {
        width: 30px; height: 30px;
        background: var(--accent-soft);
        color: var(--accent);
        border-radius: 7px;
        display: flex; align-items: center; justify-content: center;
        font-weight: 700; font-size: 0.7rem; flex-shrink: 0;
        letter-spacing: 0.05em;
    }
    .file-chip-name { color: var(--text); font-weight: 500; font-size: 0.92rem; flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
    .file-chip-size { color: var(--text-muted); font-size: 0.8rem; font-variant-numeric: tabular-nums; }

    /* ---- Stats ---- */
    .stats {
        display: grid;
        grid-template-columns: 1fr 1fr 1fr;
        gap: 0.65rem;
        margin: 0;
    }
    .stat {
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius-md);
        padding: 1rem 0.95rem;
        position: relative;
    }
    .stat .stat-lab {
        color: var(--text-muted);
        font-size: 0.72rem;
        font-weight: 600;
        margin-bottom: 0.4rem;
    }
    .stat .stat-val {
        font-size: 1.85rem;
        font-weight: 700;
        color: var(--text);
        line-height: 1;
        font-variant-numeric: tabular-nums;
        letter-spacing: -0.025em;
    }
    .stat.success { background: var(--success-soft); border-color: #A7F3D0; }
    .stat.success .stat-val { color: var(--success); }
    .stat.success .stat-lab { color: #047857; }
    .stat.warn { background: var(--warning-soft); border-color: #FDE68A; }
    .stat.warn .stat-val { color: var(--warning); }
    .stat.warn .stat-lab { color: #B45309; }
    .stat.danger { background: var(--error-soft); border-color: #FECACA; }
    .stat.danger .stat-val { color: var(--error); }
    .stat.danger .stat-lab { color: #B91C1C; }
    .stat.zero {
        opacity: 0.7;
    }

    /* ---- Buttons ---- */
    .stDownloadButton button {
        width: 100% !important;
        background: var(--accent) !important;
        color: white !important;
        border: none !important;
        border-radius: var(--radius-md) !important;
        padding: 0.85rem 1.5rem !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        font-family: 'Inter', sans-serif !important;
        letter-spacing: -0.005em !important;
        transition: all 0.15s ease !important;
        box-shadow: var(--shadow-lg) !important;
    }
    .stDownloadButton button:hover {
        background: var(--accent-hover) !important;
        transform: translateY(-1px);
        box-shadow: 0 8px 24px -6px rgba(79, 70, 229, 0.32), 0 2px 6px rgba(28, 25, 23, 0.06) !important;
    }
    .stDownloadButton button:active {
        background: var(--accent-active) !important;
        transform: translateY(0);
    }

    /* ---- Expander ---- */
    [data-testid="stExpander"] {
        border: 1px solid var(--border) !important;
        border-radius: var(--radius-md) !important;
        background: var(--surface) !important;
        margin-top: 0.85rem !important;
        box-shadow: var(--shadow-sm);
    }
    [data-testid="stExpander"] summary,
    [data-testid="stExpander"] details summary {
        font-family: 'Inter', sans-serif !important;
        font-weight: 500 !important;
        color: var(--text) !important;
        font-size: 0.88rem !important;
    }
    [data-testid="stExpander"] details summary:hover {
        color: var(--accent) !important;
    }

    /* ---- Alerts ---- */
    [data-testid="stAlert"] {
        border-radius: var(--radius-md) !important;
        border: 1px solid var(--border) !important;
        font-family: 'Inter', sans-serif !important;
        font-size: 0.88rem !important;
    }

    /* ---- DataFrame ---- */
    [data-testid="stDataFrame"] {
        border-radius: var(--radius-md);
        overflow: hidden;
        border: 1px solid var(--border);
    }

    /* ---- Spinner ---- */
    .stSpinner > div { border-top-color: var(--accent) !important; }

    /* ---- Footer ---- */
    .footer {
        text-align: center;
        margin-top: 3rem;
        padding-top: 1.5rem;
        border-top: 1px solid var(--border);
        color: var(--text-subtle);
        font-size: 0.78rem;
        letter-spacing: 0.01em;
    }
    .footer .accent { color: var(--accent); font-weight: 600; }
    .footer .sep { margin: 0 0.4rem; opacity: 0.5; }

    /* ---- Mobile ---- */
    @media (max-width: 640px) {
        .block-container { padding: 1.75rem 1rem 3rem; }
        h1.app-title { font-size: 1.55rem; }
        .stat { padding: 0.85rem 0.7rem; }
        .stat .stat-val { font-size: 1.55rem; }
        .stats { gap: 0.5rem; }
    }
</style>
""", unsafe_allow_html=True)


# ============================================================================
# Hero
# ============================================================================
st.markdown("""
<div class='hero'>
    <div class='logo-row'>
        <div class='logo-mark'></div>
        <div class='logo-text'>Conversie comenzi lentile</div>
    </div>
    <h1 class='app-title'>Arrise → Neomanager</h1>
    <p class='app-subtitle'>
        Convertește automat tabelul exportat din <span class='pill'>Arrise</span>
        în formatul cerut de <span class='pill'>Neomanager</span>.
    </p>
</div>
""", unsafe_allow_html=True)


# ============================================================================
# Step 01 — Încarcă
# ============================================================================
st.markdown("""
<div class='section-label'>
    <span class='section-num'>01</span>
    <span class='section-title'>Încarcă fișierul Arrise</span>
    <span class='section-line'></span>
</div>
""", unsafe_allow_html=True)

uploaded = st.file_uploader(
    "Tabel Arrise (.xlsx)",
    type=["xlsx", "xlsm"],
    label_visibility="collapsed",
    help="Tabel exportat din Arrise — coloane: Id · Status · Nume · Tip lentile · OD · OS · Livrare",
)


# ============================================================================
# Empty state
# ============================================================================
if uploaded is None:
    sferic_pills = "".join(
        f"<span class='pill-item'>{name.title()}</span>" for name in SFERIC_ROOTS.keys()
    )
    toric_pills = "".join(
        f"<span class='pill-item toric'>{name.title()}</span>" for name in TORIC_ROOTS.keys()
    )

    st.markdown(f"""
    <div class='pills-section'>
        <div class='pills-label'>Lentile sferice</div>
        <div class='pills'>{sferic_pills}</div>
    </div>
    <div class='pills-section'>
        <div class='pills-label'>Lentile torice</div>
        <div class='pills'>{toric_pills}</div>
    </div>
    """, unsafe_allow_html=True)

    with st.expander("Reguli aplicate"):
        st.markdown("""
- **Format dioptrie sferică**: semn + 4 cifre · `SF-1.50` → `-0150`
- **Format toric**: sferă + cilindru + ax · `SF-5.00 CYL-1.25 AX 180` → `-0500125180`
- **Plan singur** (fără CYL/AX) → rândul se sare
- **Plan cu CYL/AX** la toric → sferă=`0000` (rândul rămâne)
- **AO merge** (doar Biofinity sferic): aceeași dioptrie pe ambii ochi → o linie cu `AO` și rădăcină `LBIOFSH686`
- **Biofinity XR sferic**: `+8.50..+15.50` sau `-20.00..-12.50` → `LBFXRSH386`
- **Biofinity Toric XR**: CYL > 2.25 → `LBFXRTR387`
- **Toric cu doar sferă** (fără CYL/AX) → cade pe rădăcina sferică echivalentă
        """)

    st.markdown("""
    <div class='footer'>
        <span class='accent'>◐</span><span class='sep'>·</span> Procesare automată comenzi lentile contact
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ============================================================================
# Procesare
# ============================================================================
file_size_kb = len(uploaded.getvalue()) / 1024
size_str = f"{file_size_kb:.1f} KB" if file_size_kb < 1024 else f"{file_size_kb / 1024:.1f} MB"

with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf_in:
    tf_in.write(uploaded.getvalue())
    in_path = tf_in.name

out_path = tempfile.mktemp(suffix=".xlsx")

try:
    with st.spinner("Prelucrez fișierul..."):
        stats = process_file(in_path, out_path)
except Exception as e:
    st.error(f"Eroare la procesare: {e}")
    st.stop()


# ============================================================================
# Step 02 — Rezultat
# ============================================================================
st.markdown(f"""
<div class='section-label'>
    <span class='section-num'>02</span>
    <span class='section-title'>Rezultat</span>
    <span class='section-line'></span>
</div>
<div class='file-chip'>
    <div class='file-chip-ic'>XLS</div>
    <div class='file-chip-name'>{uploaded.name}</div>
    <div class='file-chip-size'>{size_str}</div>
</div>
""", unsafe_allow_html=True)

success_class = "success" if stats['total_rows'] > 0 else "zero"
warn_class = "warn" if stats['skipped_plan'] > 0 else "zero"
err_class = "danger" if stats['errors'] else "zero"

st.markdown(f"""
<div class='stats'>
    <div class='stat {success_class}'>
        <div class='stat-lab'>Rânduri scrise</div>
        <div class='stat-val'>{stats['total_rows']}</div>
    </div>
    <div class='stat {warn_class}'>
        <div class='stat-lab'>Sărite (Plan)</div>
        <div class='stat-val'>{stats['skipped_plan']}</div>
    </div>
    <div class='stat {err_class}'>
        <div class='stat-lab'>Erori</div>
        <div class='stat-val'>{len(stats['errors'])}</div>
    </div>
</div>
""", unsafe_allow_html=True)

if stats['unknown_products']:
    st.warning("Tipuri de lentile necunoscute: " + ", ".join(f"`{p}`" for p in stats['unknown_products']))

if stats['errors']:
    with st.expander(f"Detalii erori · {len(stats['errors'])}"):
        for err in stats['errors'][:50]:
            st.markdown(f"- {err}")
        if len(stats['errors']) > 50:
            st.markdown(f"_... și încă {len(stats['errors']) - 50} erori_")

if stats['total_rows'] > 0:
    with st.expander("Previzualizare · primele 15 rânduri"):
        wb = load_workbook(out_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = rows[0]
            data = rows[1:16]
            import pandas as pd
            df = pd.DataFrame(data, columns=header)
            st.dataframe(df, use_container_width=True, hide_index=True)


# ============================================================================
# Step 03 — Descarcă
# ============================================================================
with open(out_path, "rb") as f:
    output_bytes = f.read()

input_stem = Path(uploaded.name).stem
output_name = f"{input_stem}_neomanager.xlsx"

st.markdown(f"""
<div class='section-label'>
    <span class='section-num'>03</span>
    <span class='section-title'>Descarcă</span>
    <span class='section-line'></span>
</div>
""", unsafe_allow_html=True)

st.download_button(
    label=f"Descarcă {output_name}",
    data=output_bytes,
    file_name=output_name,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


# ============================================================================
# Footer
# ============================================================================
st.markdown("""
<div class='footer'>
    <span class='accent'>◐</span><span class='sep'>·</span> Procesare automată comenzi lentile contact
</div>
""", unsafe_allow_html=True)
